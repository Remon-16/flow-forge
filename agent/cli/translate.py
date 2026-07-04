"""用例翻译 CLI — 将已生成用例的文本字段翻译为目标语言。

Case translation CLI — translate text fields (api_name, sheet_name, remark)
of generated test cases to the target language.

Usage:
    python translate_cases.py <input_dir> [--target-lang zh_CN] [--dry-run]
"""

import argparse
import json
import logging
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# 注入 shared/py 到 sys.path，使 i18n 可导入
# Inject shared/py for i18n import
_AGENT_DIR = Path(__file__).resolve().parent.parent
_SHARED = os.path.normpath(os.path.join(str(_AGENT_DIR), "..", "shared", "py"))
if os.path.isdir(_SHARED) and _SHARED not in sys.path:
    sys.path.insert(0, _SHARED)

from i18n import _, set_lang
from i18n.loader import get_language_name
from config.translate_settings import load_translate_settings, TranslateSettings
from agents.translator import CaseTranslator

logger = logging.getLogger(__name__)

# 支持的 YAML 子目录 / Supported YAML subdirectories
_YAML_SUBDIRS = ("single_cases", "biz_flows", "interfaces")


# ============================================================================
# 参数解析 / Argument parsing
# ============================================================================


def build_translate_parser() -> argparse.ArgumentParser:
    """构建翻译工具命令行解析器。Build translator CLI argument parser."""
    p = argparse.ArgumentParser(
        prog="translate_cases.py",
        description="Flow Forge — 测试用例字段翻译工具 / Test Case Field Translator",
    )
    p.add_argument(
        "input_dir",
        help="用例目录路径 / Path to case directory (e.g., output/cases/)",
    )
    p.add_argument(
        "--output", "-o",
        default="",
        help="输出根目录（默认: <input_dir>_translated）/ Output directory (default: <input_dir>_translated)",
    )
    p.add_argument(
        "--config", "-c",
        default="translate_env.yaml",
        help="翻译配置文件路径 / Translator config file path (default: translate_env.yaml)",
    )
    p.add_argument(
        "--target-lang",
        default="",
        choices=["zh_CN", "en_US"],
        help="目标语言 / Target language (default: from config)",
    )
    p.add_argument(
        "--batch-size",
        type=int,
        default=0,
        help="每批最大用例数（覆盖配置文件）/ Max cases per batch (overrides config)",
    )
    p.add_argument(
        "--no-detection",
        action="store_true",
        help="禁用已翻译检测，全量翻译 / Disable detection, translate all cases",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="预览模式，不写入文件 / Preview without writing files",
    )
    p.add_argument(
        "--log-to-output",
        action="store_true",
        help="将日志持久化到输出目录 / Persist logs to output directory",
    )
    p.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="详细控制台日志 / Verbose console logging",
    )
    return p


# ============================================================================
# 日志初始化 / Logging setup
# ============================================================================


def _setup_logging(verbose: bool, log_dir: Optional[str] = None) -> None:
    """初始化日志系统。Setup logging.

    Args:
        verbose: 是否启用 DEBUG 级别 / Enable DEBUG level.
        log_dir: 日志文件输出目录（None=不写文件）/ Log file directory.
    """
    level = logging.DEBUG if verbose else logging.INFO
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    # 控制台 handler / Console handler
    console = logging.StreamHandler(sys.stderr)
    console.setLevel(level)
    console.setFormatter(fmt)

    root = logging.getLogger()
    root.setLevel(level)
    # 清除已有 handlers，避免重复 / Clear existing handlers
    root.handlers.clear()
    root.addHandler(console)

    # 文件 handler（可选）/ File handler (optional)
    if log_dir:
        log_path = Path(log_dir) / "logs"
        log_path.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(str(log_path / "translate.log"), encoding="utf-8")
        file_handler.setLevel(level)
        file_handler.setFormatter(fmt)
        root.addHandler(file_handler)
        logger.info(_("translate.log_enabled", path=str(log_path / "translate.log")))


# ============================================================================
# 文件扫描 / File scanning
# ============================================================================


def _scan_directory(input_dir: str) -> Tuple[List[str], List[str], bool]:
    """扫描输入目录，收集 YAML 和 Excel 文件。

    Scan input directory for YAML and Excel files.

    Returns:
        (yaml_files, excel_files, has_xlsx):
          - yaml_files: discovered YAML file paths
          - excel_files: discovered Excel file paths
          - has_xlsx: whether any .xlsx exists in the input dir
    """
    input_path = Path(input_dir)
    if not input_path.exists():
        raise FileNotFoundError(_("translate.input_not_found", dir=input_dir))

    yaml_files: List[str] = []
    excel_files: List[str] = []

    # 扫描子目录中的 YAML 文件 / Scan YAML files in subdirectories
    for subdir in _YAML_SUBDIRS:
        sub_path = input_path / subdir
        if sub_path.is_dir():
            for f in sorted(sub_path.glob("*.yaml")):
                yaml_files.append(str(f))
            for f in sorted(sub_path.glob("*.yml")):
                yaml_files.append(str(f))

    # 也扫描根目录下的 YAML（兼容扁平结构）/ Also scan root-level YAML
    for f in sorted(input_path.glob("*.yaml")):
        yaml_files.append(str(f))
    for f in sorted(input_path.glob("*.yml")):
        yaml_files.append(str(f))

    # 扫描 Excel 文件 / Scan Excel files
    for f in sorted(input_path.glob("*.xlsx")):
        excel_files.append(str(f))

    has_xlsx = len(excel_files) > 0
    return yaml_files, excel_files, has_xlsx


# ============================================================================
# 用例读取 / Case reading
# ============================================================================


def _read_yaml_cases(yaml_files: List[str]) -> List[dict]:
    """从 YAML 文件列表中读取用例。

    Read test cases from a list of YAML file paths.
    Each case dict gets a _source_file key for tracking.
    """
    import yaml

    cases: List[dict] = []
    for fpath in yaml_files:
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
            if isinstance(data, dict):
                data["_source_file"] = fpath
                cases.append(data)
        except Exception as exc:
            logger.warning(_("translate.error", error=str(exc)))
    return cases


def _read_excel_cases(excel_file: str) -> List[dict]:
    """从 Excel 文件读取用例。

    Read test cases from an Excel file.
    Returns cases with _source_file and _sheet_name tracking fields.
    """
    import openpyxl

    cases: List[dict] = []
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning(_("translate.error", error=str(exc)))
        return cases

    # Sheet 1: API Definitions → interfaces
    if "API Definitions" in wb.sheetnames:
        ws = wb["API Definitions"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in ws[1]]
        for row in rows:
            case = _row_to_case_dict(headers, row)
            if case:
                case["case_type"] = "interfaces"
                case["_source_file"] = excel_file
                case["_source_sheet"] = "API Definitions"
                cases.append(case)

    # Sheet 2: Single Cases → single
    if "Single Cases" in wb.sheetnames:
        ws = wb["Single Cases"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in ws[1]]
        for row in rows:
            case = _row_to_case_dict(headers, row)
            if case:
                case["case_type"] = "single"
                case["_source_file"] = excel_file
                case["_source_sheet"] = "Single Cases"
                cases.append(case)

    # 其他 Sheet → biz flows / Other sheets → biz flows
    for sheet_name in wb.sheetnames:
        if sheet_name in ("API Definitions", "Single Cases"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [cell.value for cell in ws[1]]
        steps = []
        for row in rows:
            step = _row_to_case_dict(headers, row)
            if step:
                steps.append(step)
        if steps:
            cases.append({
                "case_type": "biz",
                "sheet_name": sheet_name,
                "steps": steps,
                "_source_file": excel_file,
                "_source_sheet": sheet_name,
            })

    wb.close()
    return cases


def _row_to_case_dict(headers: List, values: tuple) -> Optional[dict]:
    """将 Excel 行转为用例 dict。Convert Excel row to case dict.

    将 PascalCase 列名转为 snake_case，解析 JSON 字段。
    Converts PascalCase column names to snake_case, parses JSON fields.
    """
    import json as json_mod

    if not values or all(v is None for v in values):
        return None

    # PascalCase → snake_case 映射 / PascalCase → snake_case mapping
    _PASCAL_TO_SNAKE = {
        "TestID": "test_id",
        "RelevanceID": "relevance_id",
        "Tag": "tag",
        "APIName": "api_name",
        "AppName": "app_name",
        "Method": "method",
        "URL": "url",
        "RequestHead": "request_head",
        "RequestBody": "request_body",
        "StatusCode": "status_code",
        "AssertDict": "assert_dict",
        "AssertRules": "assert_rules",
        "PreProcessors": "preprocessors",
        "PostProcessors": "postprocessors",
        "Remark": "remark",
        "StepID": "step_id",
        "Inherit": "inherit",
    }

    _JSON_FIELDS = {
        "request_head", "request_body", "assert_dict", "assert_rules",
        "preprocessors", "postprocessors",
    }

    case: dict = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        snake_key = _PASCAL_TO_SNAKE.get(str(header), str(header).lower())
        val = values[i] if i < len(values) else None
        if val is None:
            continue
        if snake_key in _JSON_FIELDS and isinstance(val, str):
            try:
                case[snake_key] = json_mod.loads(val)
            except (json_mod.JSONDecodeError, TypeError):
                case[snake_key] = val
        else:
            case[snake_key] = val

    # 检查是否为空行 / Check if row is empty
    if not any(v for v in case.values()):
        return None
    return case


# ============================================================================
# 输出写入 / Output writing
# ============================================================================


def _resolve_output_dir(input_dir: str, output_arg: str) -> str:
    """确定输出目录。Resolve output directory.

    Default: input_dir + '_translated'. If exists, append _v2, _v3...
    """
    if output_arg:
        return output_arg

    base = input_dir.rstrip("/").rstrip("\\").rstrip("/")
    candidate = base + "_translated"
    if not Path(candidate).exists():
        return candidate

    suffix = 2
    while Path(f"{base}_translated_v{suffix}").exists():
        suffix += 1
    return f"{base}_translated_v{suffix}"


def _write_yaml_cases(cases: List[dict], output_dir: str) -> int:
    """将用例列表写回 YAML 文件。Write cases back as YAML files.

    Mirrors the original directory structure (single_cases/, biz_flows/, interfaces/).
    Returns number of files written.
    """
    import yaml

    count = 0
    for case in cases:
        source = case.get("_source_file", "")
        case_type = case.get("case_type", "single")

        # 构建输出路径 / Build output path
        if case_type == "single":
            subdir = "single_cases"
            filename = Path(source).name if source else f"{case.get('test_id', 'unknown')}.yaml"
        elif case_type == "biz":
            subdir = "biz_flows"
            filename = Path(source).name if source else f"{case.get('sheet_name', 'unknown')}.yaml"
        elif case_type == "interfaces":
            subdir = "interfaces"
            filename = Path(source).name if source else f"{case.get('test_id', 'unknown')}.yaml"
        else:
            continue

        out_path = Path(output_dir) / subdir / filename
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # 移除内部跟踪字段 / Remove internal tracking fields
        clean = {k: v for k, v in case.items() if not k.startswith("_")}

        with open(out_path, "w", encoding="utf-8") as f:
            yaml.safe_dump(clean, f, allow_unicode=True, sort_keys=False, default_flow_style=False)

        logger.debug(_("translate.translated", path=str(out_path)))
        count += 1

    return count


def _write_excel_output(cases: List[dict], output_dir: str, output_path: str) -> None:
    """将翻译后的用例导出为 Excel。Export translated cases as Excel.

    将 cases list 按 case_type 分组，转为 ExcelWriter 可接受的格式。
    Groups cases by case_type and converts to ExcelWriter-compatible format.
    """
    interfaces = [c for c in cases if c.get("case_type") == "interfaces"]
    single_cases = [c for c in cases if c.get("case_type") == "single"]
    biz_flows = [c for c in cases if c.get("case_type") == "biz"]

    # 清理内部字段 / Clean internal fields
    def _clean(c: dict) -> dict:
        return {k: v for k, v in c.items() if not k.startswith("_")}

    from agents.excel_writer import ExcelWriter
    ExcelWriter.write(
        [_clean(c) for c in interfaces],
        [_clean(c) for c in single_cases],
        [[_clean(s) for s in (c.get("steps", []) if isinstance(c.get("steps"), list) else [])] for c in biz_flows],
        output_path,
    )
    logger.info(_("translate.writing_excel", path=output_path))


# ============================================================================
# 主入口 / Main entry
# ============================================================================


def translate_main() -> int:
    """翻译工具主入口。Translator main entry point.

    Returns:
        0 on success, 2 on error.
    """
    parser = build_translate_parser()
    args = parser.parse_args()

    # 加载配置 / Load config
    config_path = args.config
    if not os.path.isfile(config_path):
        print(f"Error: Config file not found: {config_path}", file=sys.stderr)
        print("  Run: cp translate_env.example.yaml translate_env.yaml", file=sys.stderr)
        print("  Then edit translate_env.yaml with your settings.", file=sys.stderr)
        return 2

    try:
        settings: TranslateSettings = load_translate_settings(config_path)
    except Exception as exc:
        print(f"Error loading config: {exc}", file=sys.stderr)
        return 2

    # 覆盖参数 / Override from CLI
    if args.target_lang:
        settings.target_lang = args.target_lang
    if args.batch_size:
        settings.batch_size = args.batch_size
    if args.no_detection:
        settings.detection_enabled = False

    # 初始化 i18n 和日志 / Init i18n and logging
    set_lang(settings.target_lang)
    _setup_logging(args.verbose)

    # 确定输出目录 / Determine output directory
    output_dir = _resolve_output_dir(args.input_dir, args.output)
    if args.log_to_output or settings.log_to_output:
        # 重新初始化日志以包含文件输出 / Re-init logging with file output
        _setup_logging(args.verbose, log_dir=output_dir)

    # 检查 API key / Check API key
    if not settings.llm_api_key:
        logger.error(_("translate.no_api_key"))
        return 2

    # 扫描目录 / Scan directory
    logger.info(_("translate.scanning"))
    try:
        yaml_files, excel_files, has_xlsx = _scan_directory(args.input_dir)
    except FileNotFoundError as exc:
        logger.error(str(exc))
        return 2

    # 判断输入源 / Determine input source
    if yaml_files:
        logger.info(_("translate.yaml_priority"))
        logger.info(_("translate.found_files", count=len(yaml_files)))
        cases = _read_yaml_cases(yaml_files)
        output_format = "yaml"
    elif excel_files:
        logger.info(_("translate.excel_only"))
        logger.info(_("translate.found_files", count=len(excel_files)))
        cases = []
        for xf in excel_files:
            cases.extend(_read_excel_cases(xf))
        output_format = "excel"
    else:
        logger.error(_("translate.no_files"))
        return 2

    if not cases:
        logger.error(_("translate.no_files"))
        return 2

    # 已翻译检测 / Already-translated detection
    translated_count = 0
    skipped_cases = []
    if settings.detection_enabled:
        to_translate = []
        for case in cases:
            needs = CaseTranslator.case_needs_translation(
                case, settings.target_lang, settings.cjk_threshold
            )
            if needs:
                to_translate.append(case)
            else:
                skipped_cases.append(case)
        skipped = len(skipped_cases)
        if skipped > 0:
            logger.info(_("translate.detection_skip", count=skipped))
        cases = to_translate

    if not cases:
        logger.info(_(
            "translate.summary",
            total=len(skipped_cases), translated=0,
            skipped=len(skipped_cases), failed=0,
            output_dir=output_dir,
        ))
        return 0

    # 分批 / Batching
    batch_size = settings.batch_size if settings.batch_size > 0 else len(cases)
    batches = [cases[i:i + batch_size] for i in range(0, len(cases), batch_size)]
    total_batches = len(batches)
    logger.info(_("translate.batching", batches=total_batches, size=batch_size))

    # 创建翻译智能体 / Create translator agent
    translator = CaseTranslator(settings)
    logger.info(_(
        "translate.translating",
        model=settings.llm_model,
    ).replace("{model}", settings.llm_model).replace("  → CaseTranslator 正在调用 LLM ({model})...", f"  → CaseTranslator 正在调用 LLM..."))

    # 逐批翻译 / Translate batch by batch
    translated_cases = []
    failed_count = 0
    for i, batch in enumerate(batches, 1):
        logger.info(_("translate.batch_progress", current=i, total=total_batches, count=len(batch)))
        try:
            result = translator.translate_batch(batch)
            translated_cases.extend(result)
            logger.info(_("translate.batch_done", current=i))
        except Exception as exc:
            logger.error(_("translate.error", error=str(exc)))
            failed_count += len(batch)
            translated_cases.extend(batch)  # 保留原始数据 / Keep original

    # 合并跳过的用例 / Merge skipped cases
    all_output = translated_cases + skipped_cases

    # 写入输出 / Write output
    if not args.dry_run:
        if output_format == "yaml":
            written = _write_yaml_cases(all_output, output_dir)
            # 若输入含 Excel，也导出 Excel / Also export Excel if input had it
            if has_xlsx:
                excel_path = str(Path(output_dir) / "test_cases.xlsx")
                _write_excel_output(all_output, output_dir, excel_path)
        else:
            excel_path = str(Path(output_dir) / "test_cases.xlsx")
            _write_excel_output(all_output, output_dir, excel_path)

    # 汇总 / Summary
    total = len(all_output)
    logger.info(_(
        "translate.summary",
        total=total,
        translated=len(translated_cases),
        skipped=len(skipped_cases),
        failed=failed_count,
        output_dir=output_dir,
    ))

    if args.dry_run:
        logger.info(_("translate.dry_run"))

    return 2 if failed_count > 0 else 0
