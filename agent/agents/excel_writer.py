"""ExcelWriter: write structured test cases to .xlsx matching executor format."""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models.schema import BizFlow, InterfaceDef, SingleTestCase

logger = logging.getLogger(__name__)

# Column headers matching executor's ExcelParser expectations
_API_COLUMNS = [
    "TestID", "APIName", "AppName", "Method", "URL",
    "RequestHead", "RequestBody", "StatusCode", "AssertDict", "AssertRules",
    "PreProcessors", "PostProcessors", "Remark",
]

_CASE_COLUMNS = [
    "TestID", "RelevanceID", "Tag",
    "APIName", "AppName", "Method", "URL",
    "RequestHead", "RequestBody", "StatusCode", "AssertDict", "AssertRules",
    "PreProcessors", "PostProcessors", "Remark",
]

_BIZ_COLUMNS = [
    "StepID", "RelevanceID", "Trans",
    "APIName", "AppName", "Method", "URL",
    "RequestHead", "RequestBody", "StatusCode", "AssertDict", "AssertRules",
    "PreProcessors", "PostProcessors", "Tag", "Remark",
]

_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelWriter:
    """Write generated test cases to Excel format matching the executor."""

    @staticmethod
    def yaml_to_excel(output_dir: str, output_path: str) -> str:
        """Convert YAML files in output_dir to a single Excel workbook.

        Reads interfaces/, single_cases/, biz_flows/ subdirectories.
        """
        from writers.yaml_writer import YamlWriter

        interfaces = YamlWriter.read_interfaces(output_dir)
        single_cases = YamlWriter.read_single_cases(output_dir)
        biz_flows = YamlWriter.read_biz_flows(output_dir)

        return ExcelWriter.write(interfaces, single_cases, biz_flows, output_path)

    @staticmethod
    def write(
        interfaces: List[Any],
        single_cases: List[Any],
        biz_flows: List[Any],
        output_path: str,
    ) -> str:
        """Write all test cases to Excel file.

        Sheet 1: API Definitions
        Sheet 2: Single Test Cases
        Sheets 3+: One sheet per BizFlow

        Accepts both dataclass instances and plain dicts.
        Returns the output path.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: API Definitions
        ws1 = wb.active
        ws1.title = "API Definitions"
        ExcelWriter._write_api_sheet(ws1, interfaces)

        # Sheet 2: Single Cases
        ws2 = wb.create_sheet("Single Cases")
        if single_cases:
            ExcelWriter._write_single_sheet(ws2, single_cases)
        else:
            ExcelWriter._write_headers(ws2, _CASE_COLUMNS)

        # Sheets 3+: Biz Flows
        for flow in biz_flows:
            safe_name = ExcelWriter._safe_sheet_name(
                ExcelWriter._get_attr(flow, "sheet_name", "BizFlow")
            )
            ws = wb.create_sheet(safe_name)
            ExcelWriter._write_biz_sheet(ws, flow)

        # Ensure output directory exists
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(str(out_path))
        logger.info(
            "Excel written to %s (%d interfaces, %d single cases, %d biz flows)",
            output_path, len(interfaces), len(single_cases), len(biz_flows),
        )
        return str(out_path)

    @staticmethod
    def _get_attr(obj: Any, name: str, default: Any = "") -> Any:
        """Get attribute from dataclass or key from dict."""
        if isinstance(obj, dict):
            return obj.get(name, default)
        return getattr(obj, name, default)

    @staticmethod
    def _write_api_sheet(ws, interfaces: List[Any]) -> None:
        ExcelWriter._write_headers(ws, _API_COLUMNS)
        for row_idx, iface in enumerate(interfaces, start=2):
            g = lambda n, d="": ExcelWriter._get_attr(iface, n, d)
            ExcelWriter._write_row(ws, row_idx, [
                g("test_id"),
                g("api_name"),
                g("app_name"),
                g("method"),
                g("url"),
                json.dumps(g("request_head"), ensure_ascii=False) if g("request_head") else "",
                json.dumps(g("request_body"), ensure_ascii=False) if g("request_body") else "",
                g("status_code", 200),
                json.dumps(g("assert_dict"), ensure_ascii=False) if g("assert_dict") else "",
                json.dumps(g("assert_rules"), ensure_ascii=False) if g("assert_rules") else "",
                json.dumps(g("preprocessors"), ensure_ascii=False) if g("preprocessors") else "",
                json.dumps(g("postprocessors"), ensure_ascii=False) if g("postprocessors") else "",
                g("remark"),
            ])

    @staticmethod
    def _write_single_sheet(ws, cases: List[Any]) -> None:
        ExcelWriter._write_headers(ws, _CASE_COLUMNS)
        for row_idx, case in enumerate(cases, start=2):
            g = lambda n, d="": ExcelWriter._get_attr(case, n, d)
            ExcelWriter._write_row(ws, row_idx, [
                g("test_id"),
                g("relevance_id"),
                g("tag", "P1"),
                g("api_name"),
                g("app_name"),
                g("method"),
                g("url"),
                json.dumps(g("request_head"), ensure_ascii=False) if g("request_head") else "",
                json.dumps(g("request_body"), ensure_ascii=False) if g("request_body") else "",
                g("status_code", 200),
                json.dumps(g("assert_dict"), ensure_ascii=False) if g("assert_dict") else "",
                json.dumps(g("assert_rules"), ensure_ascii=False) if g("assert_rules") else "",
                json.dumps(g("preprocessors"), ensure_ascii=False) if g("preprocessors") else "",
                json.dumps(g("postprocessors"), ensure_ascii=False) if g("postprocessors") else "",
                g("remark"),
            ])

    @staticmethod
    def _write_biz_sheet(ws, flow: Any) -> None:
        ExcelWriter._write_headers(ws, _BIZ_COLUMNS)
        steps = ExcelWriter._get_attr(flow, "steps", [])
        for row_idx, step in enumerate(steps, start=2):
            g = lambda n, d="": ExcelWriter._get_attr(step, n, d)
            ExcelWriter._write_row(ws, row_idx, [
                g("step_id"),
                g("relevance_id"),
                g("trans"),
                g("api_name"),
                g("app_name"),
                g("method"),
                g("url"),
                json.dumps(g("request_head"), ensure_ascii=False) if g("request_head") else "",
                json.dumps(g("request_body"), ensure_ascii=False) if g("request_body") else "",
                g("status_code", 200),
                json.dumps(g("assert_dict"), ensure_ascii=False) if g("assert_dict") else "",
                json.dumps(g("assert_rules"), ensure_ascii=False) if g("assert_rules") else "",
                json.dumps(g("preprocessors"), ensure_ascii=False) if g("preprocessors") else "",
                json.dumps(g("postprocessors"), ensure_ascii=False) if g("postprocessors") else "",
                g("tag", "P1"),
                g("remark"),
            ])

    @staticmethod
    def _write_headers(ws, columns: List[str]) -> None:
        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER

    @staticmethod
    def _write_row(ws, row_idx: int, values: list) -> None:
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            cell.font = Font(name="微软雅黑", size=10)

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Sanitize sheet name to Excel's 31-char limit and invalid chars."""
        # Replace invalid chars
        safe = name.replace(":", "-").replace("\\", "-").replace("/", "-")
        safe = safe.replace("*", "-").replace("?", "-").replace("[", "").replace("]", "")
        return safe[:31]

    @staticmethod
    def _auto_width(ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
