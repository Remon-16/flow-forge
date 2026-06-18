import json
import logging
import os
import re
from typing import Any, Dict, List, Optional

import openpyxl


logger = logging.getLogger(__name__)

_CASE_SHEET_INDEX = 1
_BIZ_START_SHEET_INDEX = 2

_SIMPLE_FIELDS = ("APIName", "AppName", "Method", "URL", "StatusCode")
_JSON_FIELDS = ("RequestHead", "RequestBody")
_SHEET2_REQUIRED = ("TestID", "RelevanceID")
_BIZ_REQUIRED = ("StepID", "RelevanceID")

_CHINESE_RE = re.compile(r"[一-鿿㐀-䶿豈-﫿]")


class ExcelParseError(Exception):
    """Raised when Excel cell content fails validation."""


class StepIDConflictError(Exception):
    """Raised when duplicate StepID values exist within a biz flow sheet."""


class ExcelParser:
    """Reads a multi-sheet Excel test-case file and returns merged test cases."""

    def __init__(self, file_path: str):
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        self.file_path = file_path

    def parse(self, api_mode: str = "single") -> Dict[str, Any]:
        """Parse Excel based on apiMode.

        Returns: {"single_cases": [...], "biz_flows": [...]}
          biz_flows: [{"sheet_name": str, "steps": [...], "parse_error": str|None}]
        """
        result: Dict[str, Any] = {"single_cases": [], "biz_flows": []}

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        try:
            if len(wb.sheetnames) < 2:
                raise ValueError(
                    f"Expected at least 2 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
                )

            if api_mode in ("single", "all"):
                test_cases = self._read_sheet_rows(
                    wb.worksheets[_CASE_SHEET_INDEX], list(_SHEET2_REQUIRED)
                )
                logger.info("Read %d single test cases from sheet 2", len(test_cases))
                result["single_cases"] = self._merge_cases(test_cases)
                logger.info("Merged into %d single test cases", len(result["single_cases"]))

            if api_mode in ("biz", "all"):
                for idx in range(_BIZ_START_SHEET_INDEX, len(wb.sheetnames)):
                    ws = wb.worksheets[idx]
                    sheet_name = ws.title
                    try:
                        biz_flow = self._parse_biz_flow(ws, sheet_name)
                        result["biz_flows"].append(biz_flow)
                    except (ExcelParseError, StepIDConflictError) as e:
                        logger.error("Biz flow '%s' parse error: %s", sheet_name, e)
                        result["biz_flows"].append({
                            "sheet_name": sheet_name,
                            "steps": [],
                            "parse_error": str(e),
                        })

                logger.info("Parsed %d biz flows", len(result["biz_flows"]))

            return result
        finally:
            wb.close()

    def _read_sheet_rows(
        self, ws, required_columns: List[str]
    ) -> List[Dict[str, Any]]:
        headers = []
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                headers.append((col_idx, str(cell.value).strip()))

        header_names = [h[1] for h in headers]
        missing = [c for c in required_columns if c not in header_names]
        if missing:
            raise ValueError(
                f"Sheet '{ws.title}' is missing required columns: {missing}"
            )

        col_map = {h[1]: h[0] for h in headers}
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for name, col in col_map.items():
                row_data[name] = ws.cell(row=row_idx, column=col).value
            if all(v is None for v in row_data.values()):
                continue
            rows.append(row_data)

        return rows

    def _parse_biz_flow(
        self, ws, sheet_name: str
    ) -> Dict[str, Any]:
        raw_steps = self._read_sheet_rows(ws, list(_BIZ_REQUIRED))

        step_ids = []
        for row in raw_steps:
            sid = str(row.get("StepID", ""))
            if sid:
                step_ids.append(sid)

        dupes = {sid for sid in step_ids if step_ids.count(sid) > 1}
        if dupes:
            raise StepIDConflictError(
                f"Duplicate StepID in biz flow '{sheet_name}': {dupes}"
            )

        for row in raw_steps:
            trans_val = row.get("Trans")
            if trans_val not in (None, ""):
                self._validate_trans(str(trans_val), str(row.get("StepID", "")), sheet_name)

        merged_steps = self._merge_cases(raw_steps, is_biz=True)
        return {"sheet_name": sheet_name, "steps": merged_steps, "parse_error": None}

    def _validate_trans(self, trans_str: str, step_id: str, sheet_name: str) -> None:
        stripped = trans_str.strip()
        if not stripped:
            return

        if _CHINESE_RE.search(stripped):
            raise ExcelParseError(
                f"Trans field contains Chinese characters in sheet '{sheet_name}', "
                f"StepID='{step_id}': {stripped}"
            )

        pairs = [p.strip() for p in stripped.split(",")]
        for pair in pairs:
            if not pair:
                continue
            if "=" not in pair:
                raise ExcelParseError(
                    f"Trans field format error (expected key=value) in sheet '{sheet_name}', "
                    f"StepID='{step_id}': '{pair}'"
                )
            key, value = pair.split("=", 1)
            key = key.strip()
            value = value.strip()
            if not key:
                raise ExcelParseError(
                    f"Trans field has empty key in sheet '{sheet_name}', "
                    f"StepID='{step_id}': '{pair}'"
                )
            if not value:
                raise ExcelParseError(
                    f"Trans field has empty value for key '{key}' in sheet '{sheet_name}', "
                    f"StepID='{step_id}'"
                )

            open_brackets = value.count("[") - value.count("]")
            if open_brackets != 0:
                raise ExcelParseError(
                    f"Trans field has mismatched brackets '[' ']' in sheet '{sheet_name}', "
                    f"StepID='{step_id}', key='{key}': {value}"
                )
            open_parens = value.count("(") - value.count(")")
            if open_parens != 0:
                raise ExcelParseError(
                    f"Trans field has mismatched brackets '(' ')' in sheet '{sheet_name}', "
                    f"StepID='{step_id}', key='{key}': {value}"
                )

    def _merge_cases(
        self,
        test_cases: List[Dict[str, Any]],
        is_biz: bool = False,
    ) -> List[Dict[str, Any]]:
        merged = []
        for tc in test_cases:
            merged.append(self._build_result(tc, is_biz))
        return merged

    def _build_result(
        self,
        tc: Dict[str, Any],
        is_biz: bool = False,
    ) -> Dict[str, Any]:
        result = {}

        for field in _SIMPLE_FIELDS:
            result[self._normalize_key(field)] = tc.get(field)

        for field in _JSON_FIELDS:
            result[self._normalize_key(field)] = self._safe_parse_json(
                tc.get(field), field, tc.get("TestID") or tc.get("StepID")
            )

        assert_field = tc.get("AssertDict")
        if assert_field not in (None, ""):
            result["assert_dict"] = self._safe_parse_json(
                assert_field, "AssertDict", tc.get("TestID") or tc.get("StepID")
            )
        else:
            result["assert_dict"] = {}

        assert_rules_field = tc.get("AssertRules")
        if assert_rules_field not in (None, ""):
            parsed = self._safe_parse_json(
                assert_rules_field, "AssertRules", tc.get("TestID") or tc.get("StepID")
            )
            result["assert_rules"] = parsed if isinstance(parsed, list) else []
        else:
            result["assert_rules"] = []

        if is_biz:
            result["test_id"] = str(tc.get("StepID", ""))
            result["step_id"] = str(tc.get("StepID", ""))
            trans_raw = tc.get("Trans")
            result["trans"] = str(trans_raw).strip() if trans_raw not in (None, "") else ""
        else:
            result["test_id"] = str(tc.get("TestID", ""))

        result["tag"] = str(tc.get("Tag", "")) if tc.get("Tag") is not None else ""
        result["remark"] = str(tc.get("Remark", "")) if tc.get("Remark") is not None else ""

        # PreProcessors / PostProcessors
        for col, key in [("PreProcessors", "preprocessors"), ("PostProcessors", "postprocessors")]:
            raw = tc.get(col)
            if raw not in (None, ""):
                parsed = self._safe_parse_json(
                    raw, col, tc.get("TestID") or tc.get("StepID")
                )
                result[key] = parsed if isinstance(parsed, list) else []
            else:
                result[key] = []

        return result

    @staticmethod
    def _normalize_key(field: str) -> str:
        mapping = {
            "APIName": "api_name",
            "AppName": "app_name",
            "Method": "method",
            "URL": "url",
            "StatusCode": "status_code",
            "RequestHead": "request_head",
            "RequestBody": "request_body",
        }
        return mapping.get(field, field.lower())

    @staticmethod
    def excel_str_to_dict(s: str) -> Dict[str, Any]:
        s = s.replace('“', '"')
        s = s.replace('”', '"')
        s = s.replace('‘', '"')
        s = s.replace('’', '"')
        s = s.replace("'", '"')
        return json.loads(s)

    @staticmethod
    def _safe_parse_json(
        raw: Any, field_name: str, test_id: Optional[str]
    ) -> Dict[str, Any]:
        if raw is None:
            return {}
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, str):
            stripped = raw.strip()
            if not stripped:
                return {}
            try:
                return ExcelParser.excel_str_to_dict(stripped)
            except (json.JSONDecodeError, ValueError):
                logger.warning(
                    "Failed to parse %s as JSON for '%s', using empty dict",
                    field_name,
                    test_id,
                )
                return {}
        return {}
