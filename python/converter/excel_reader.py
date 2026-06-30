"""Read an Excel workbook and extract structured test case data.

Returns data in snake_case dicts suitable for writing to YAML.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

from .field_mapping import convert_row_to_snake
from .common.columns import API_COLUMNS, CASE_COLUMNS, BIZ_COLUMNS

logger = logging.getLogger(__name__)

_API_SHEET = "API Definitions"
_CASE_SHEET = "Single Cases"


def _parse_excel_sheet(
    ws: Any, headers: list[str]
) -> list[dict[str, object]]:
    """Read rows from a worksheet and build dicts keyed by header name."""
    rows: list[dict[str, object]] = []
    for row_idx in range(2, ws.max_row + 1):
        row_data: dict[str, object] = {}
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                row_data[header] = value
        if row_data:  # skip fully empty rows
            rows.append(row_data)
    return rows


def read_excel(file_path: str) -> dict[str, Any]:
    """Read an Excel file and return structured test case data.

    Returns:
        {
            "interfaces": [...],   # snake_case dicts
            "single_cases": [...], # snake_case dicts
            "biz_flows": [         # each with "sheet_name" + "steps"
                {"sheet_name": "...", "steps": [...]},
                ...
            ],
        }
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    interfaces: list[dict[str, object]] = []
    single_cases: list[dict[str, object]] = []
    biz_flows: list[dict[str, object]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not ws or ws.max_row < 2:
            continue

        if sheet_name == _API_SHEET:
            raw_rows = _parse_excel_sheet(ws, API_COLUMNS)
            for row in raw_rows:
                converted = convert_row_to_snake(row, parse_json=True)
                if converted:
                    interfaces.append(converted)
            logger.info("Read %d interface definitions from sheet '%s'", len(interfaces), sheet_name)

        elif sheet_name == _CASE_SHEET:
            raw_rows = _parse_excel_sheet(ws, CASE_COLUMNS)
            for row in raw_rows:
                converted = convert_row_to_snake(row, parse_json=True)
                if converted:
                    single_cases.append(converted)
            logger.info("Read %d single cases from sheet '%s'", len(single_cases), sheet_name)

        else:
            # Any other sheet is treated as a biz flow
            raw_rows = _parse_excel_sheet(ws, BIZ_COLUMNS)
            steps: list[dict[str, object]] = []
            for row in raw_rows:
                converted = convert_row_to_snake(row, parse_json=True)
                if converted:
                    steps.append(converted)
            if steps:
                biz_flows.append({
                    "sheet_name": sheet_name,
                    "steps": steps,
                })
            logger.info("Read biz flow '%s' with %d steps", sheet_name, len(steps))

    wb.close()
    return {
        "interfaces": interfaces,
        "single_cases": single_cases,
        "biz_flows": biz_flows,
    }
