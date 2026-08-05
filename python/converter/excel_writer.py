"""Write structured test case data to an Excel workbook.

Produces the same multi-sheet format that the Agent's ExcelWriter and
the Python executor's ExcelParser expect.
"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .common.columns import API_COLUMNS, CASE_COLUMNS, BIZ_COLUMNS
from .common.utils import safe_sheet_name
from i18n import _

logger = logging.getLogger(__name__)


def _get_font_name() -> str:
    """返回配置的 Excel 字体名。优先读取环境变量 EXCEL_FONT，其次尝试 config_manager，默认微软雅黑。

    Return the configured Excel font name. Checks EXCEL_FONT env var first,
    then config_manager if initialized, falling back to Microsoft YaHei.
    """
    env_val = os.environ.get("EXCEL_FONT", "").strip()
    if env_val:
        return env_val
    try:
        from config.config_manager import get, is_initialized
        if is_initialized():
            return get("excel_font", "微软雅黑")
    except (ImportError, RuntimeError):
        pass
    return "微软雅黑"

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_headers(ws: Any, columns: list[str]) -> None:
    header_font = Font(name=_get_font_name(), bold=True, size=11, color="FFFFFF")
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _write_row(ws: Any, row_idx: int, values: list[object]) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _THIN_BORDER
        cell.font = Font(name=_get_font_name(), size=10)


def _get(obj: dict[str, object] | Any, key: str, default: object = "") -> object:
    """Get a value from a dict by key, falling back to default."""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _serialize_json(value: object) -> str:
    """Serialize a value to a JSON string if it's not already a string."""
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def write_excel(
    interfaces: list[dict[str, object]] | None,
    single_cases: list[dict[str, object]] | None,
    biz_flows: list[dict[str, object]] | None,
    output_path: str,
) -> str:
    """Write test case data to an Excel workbook.

    Args:
        interfaces: List of interface definition dicts (snake_case keys).
        single_cases: List of single API case dicts (snake_case keys).
        biz_flows: List of biz flow dicts, each with ``sheet_name`` and ``steps``.
        output_path: Destination .xlsx file path.

    Returns:
        The resolved output path.
    """
    wb = openpyxl.Workbook()

    # Sheet 1: API Definitions
    ws1 = wb.active
    ws1.title = "API Definitions"
    _write_headers(ws1, API_COLUMNS)
    if interfaces:
        for row_idx, iface in enumerate(interfaces, start=2):
            _write_row(ws1, row_idx, [
                _get(iface, "test_id"),
                _get(iface, "api_name"),
                _get(iface, "app_name"),
                _get(iface, "method"),
                _get(iface, "url"),
                _serialize_json(_get(iface, "request_head")),
                _serialize_json(_get(iface, "request_body")),
                _get(iface, "status_code", 200),
                _serialize_json(_get(iface, "assert_dict")),
                _serialize_json(_get(iface, "assert_rules")),
                _serialize_json(_get(iface, "preprocessors")),
                _serialize_json(_get(iface, "postprocessors")),
                _get(iface, "remark"),
            ])

    # Sheet 2: Single Cases
    ws2 = wb.create_sheet("Single Cases")
    _write_headers(ws2, CASE_COLUMNS)
    if single_cases:
        for row_idx, case in enumerate(single_cases, start=2):
            _write_row(ws2, row_idx, [
                _get(case, "test_id"),
                _get(case, "relevance_id"),
                _get(case, "tag", "P1"),
                _get(case, "api_name"),
                _get(case, "app_name"),
                _get(case, "method"),
                _get(case, "url"),
                _serialize_json(_get(case, "request_head")),
                _serialize_json(_get(case, "request_body")),
                _get(case, "status_code", 200),
                _serialize_json(_get(case, "assert_dict")),
                _serialize_json(_get(case, "assert_rules")),
                _serialize_json(_get(case, "preprocessors")),
                _serialize_json(_get(case, "postprocessors")),
                _get(case, "remark"),
            ])

    # Sheets 3+: Biz Flows
    if biz_flows:
        for flow in biz_flows:
            sheet_name = str(_get(flow, "sheet_name", "BizFlow"))
            safe_name = safe_sheet_name(sheet_name)
            ws = wb.create_sheet(safe_name)
            _write_headers(ws, BIZ_COLUMNS)
            steps = _get(flow, "steps", [])
            if isinstance(steps, list):
                for row_idx, step in enumerate(steps, start=2):
                    _write_row(ws, row_idx, [
                        _get(step, "step_id"),
                        _get(step, "relevance_id"),
                        _serialize_json(_get(step, "inherit")),
                        _get(step, "api_name"),
                        _get(step, "app_name"),
                        _get(step, "method"),
                        _get(step, "url"),
                        _serialize_json(_get(step, "request_head")),
                        _serialize_json(_get(step, "request_body")),
                        _get(step, "status_code", 200),
                        _serialize_json(_get(step, "assert_dict")),
                        _serialize_json(_get(step, "assert_rules")),
                        _serialize_json(_get(step, "preprocessors")),
                        _serialize_json(_get(step, "postprocessors")),
                        _get(step, "tag", "P1"),
                        _get(step, "remark"),
                    ])

    # Ensure output directory exists
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    logger.info(
        _(
            "converter.excel_written_summary",
            path=output_path,
            interfaces=len(interfaces) if interfaces else 0,
            single=len(single_cases) if single_cases else 0,
            biz=len(biz_flows) if biz_flows else 0,
        )
    )
    return str(out_path)
